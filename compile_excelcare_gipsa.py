import openpyxl
import os
import json
import sys

def compile_excelcare_gipsa():
    filepath = r"S:/Sid Work/1. Apollo/@ Apollo Guwahti/Tarriff Working/Tarrif Reporting Format/Excelcare/SOC's from Dipjyoti/APL EXL HOSP_SOC.xlsx"
    output_path = r"S:\Sid Work\1. Apollo\@ Apollo Guwahti\Tarriff Working\Tarrif Reporting Format\Tarrif Masterss\excelcare_gipsa_2026.json"
    
    if not os.path.exists(filepath):
        print(f"Warning: File not found at {filepath}")
        return
        
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    unique_gipsa = {}
    
    # 1. Parse Bed Charges sheet (corresponds to Sheet 1)
    if "Bed Charges" in wb.sheetnames:
        sheet1 = wb["Bed Charges"]
    else:
        sheet1 = wb.worksheets[0]
        
    for row in sheet1.iter_rows(values_only=True):
        if len(row) < 7:
            continue
        # In Bed Charges, code is at index 2 (Column C), name at index 3 (Column D), rate at index 6 (Column G)
        code = str(row[2]).strip() if row[2] is not None else ""
        name = str(row[3]).strip() if row[3] is not None else ""
        if not code or not code.split('.')[0].isdigit():
            continue
        try:
            code_val = float(code)
            if code_val < 1:
                continue
        except ValueError:
            continue
        if not name:
            continue
            
        rate_val = row[6]
        rate_num = 0.0
        if rate_val is not None:
            try:
                rate_num = float(rate_val)
            except ValueError:
                pass
                
        if rate_num > 0:
            unique_gipsa[code] = {
                "id": code,
                "name": name,
                "rate": rate_num
            }
            
    # 2. Parse Other Tariff sheet (corresponds to Sheet 4)
    if "Other Tariff" in wb.sheetnames:
        sheet2 = wb["Other Tariff"]
    else:
        sheet2 = wb.worksheets[3]
        
    for row in sheet2.iter_rows(values_only=True):
        if len(row) < 6:
            continue
        # In Other Tariff, code is at index 4 (Column E), name at index 5 (Column F)
        code = str(row[4]).strip() if row[4] is not None else ""
        name = str(row[5]).strip() if row[5] is not None else ""
        
        if not code or not code.split('.')[0].isdigit():
            continue
        try:
            code_val = float(code)
            if code_val < 1:
                continue
        except ValueError:
            continue
        if not name:
            continue
            
        rate_num = 0.0
        # OPD rate is at index 7 (Column H)
        if len(row) > 7 and row[7] is not None:
            try:
                rate_num = float(row[7])
            except ValueError:
                pass
                
        if rate_num <= 0:
            # Scan remaining columns (from index 7 onwards) for first positive rate
            for c_idx in range(7, len(row)):
                val = row[c_idx]
                if val is not None:
                    try:
                        val_num = float(val)
                        if val_num > 0:
                            rate_num = val_num
                            break
                    except ValueError:
                        pass
                        
        if code not in unique_gipsa or (unique_gipsa[code]["rate"] == 0 and rate_num > 0):
            unique_gipsa[code] = {
                "id": code,
                "name": name,
                "rate": rate_num
            }
            
    results = list(unique_gipsa.values())
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2)
    print(f"Excelcare GIPSA 2026 compiled successfully. Total records: {len(results)}")

if __name__ == "__main__":
    compile_excelcare_gipsa()
