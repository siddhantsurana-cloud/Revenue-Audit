import openpyxl
import os
import json
import hashlib
import sys

def get_file_hash(filepath):
    if not os.path.exists(filepath):
        return ""
    hasher = hashlib.md5()
    try:
        with open(filepath, 'rb') as f:
            buf = f.read(65536)
            while len(buf) > 0:
                hasher.update(buf)
                buf = f.read(65536)
    except Exception:
        return ""
    return hasher.hexdigest()

def clean_rate(val):
    if val is None:
        return 0.0
    val_str = str(val).replace(",", "").replace(" ", "").strip()
    try:
        return float(val_str)
    except ValueError:
        return 0.0

def parse_template(filepath, unit_name):
    if not os.path.exists(filepath):
        print(f"Warning: File not found at {filepath}")
        return {}
        
    print(f"Parsing {unit_name} template: {os.path.basename(filepath)}...")
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    compiled = {} # code -> {name, rate, dept}
    
    for sheetname in wb.sheetnames:
        # Skip administrative or text-heavy pages
        if sheetname.lower() in ["cover page", "sheet1", "sheet2", "others", "group docs name"]:
            continue
            
        sheet = wb[sheetname]
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
            
        # Detect headers dynamically
        code_col = None
        name_col = None
        rate_col = None
        header_row_idx = -1
        
        for r_idx, r in enumerate(rows[:20]): # scan first 20 rows for headers
            for c_idx, val in enumerate(r):
                if val is None:
                    continue
                val_upper = str(val).strip().upper()
                if any(x in val_upper for x in ["SERVICE ID", "NEW CODE", "CODE"]) and not any(x in val_upper for x in ["TYPE", "NAME"]):
                    code_col = c_idx
                elif any(x in val_upper for x in ["SERVICE NAME", "SERVICE", "PARTICULARS", "BED CATEGORY", "ROOM TARIFF"]):
                    name_col = c_idx
                elif any(x in val_upper for x in ["AMOUNT", "TARIFF", "RATE", "CHARGE", "(AMT IN RS)"]):
                    if "GST" not in val_upper and "DISCOUNT" not in val_upper and "PERC" not in val_upper:
                        rate_col = c_idx
            
            if code_col is not None and name_col is not None and rate_col is not None:
                header_row_idx = r_idx
                break
                
        # If we failed to find code, name, and rate, try looser matching
        if header_row_idx == -1:
            code_col, name_col, rate_col = None, None, None
            for r_idx, r in enumerate(rows[:20]):
                for c_idx, val in enumerate(r):
                    if val is None:
                        continue
                    val_upper = str(val).strip().upper()
                    if "CODE" in val_upper:
                        code_col = c_idx
                    elif any(x in val_upper for x in ["SERVICE", "PARTICULARS"]):
                        name_col = c_idx
                    elif any(x in val_upper for x in ["AMOUNT", "RATE", "TARIFF"]):
                        rate_col = c_idx
                if name_col is not None and rate_col is not None:
                    header_row_idx = r_idx
                    break
                    
        if header_row_idx == -1:
            # Fallback to default columns if sheet name indicates structured billing items
            if sheetname.lower() in ["lab", "radiology", "cardiology", "neurology", "investigations", "nephrology"]:
                code_col = 1
                name_col = 2
                rate_col = 3
                header_row_idx = 3
            else:
                continue
                
        if code_col is None:
            # Skip sheet if there is no code column
            continue
            
        # Parse data rows
        for r in rows[header_row_idx + 1:]:
            if len(r) <= max(code_col, name_col, rate_col):
                continue
                
            code_val = r[code_col]
            name_val = r[name_col]
            rate_val = r[rate_col]
            
            code = str(code_val).strip() if code_val is not None else ""
            name = str(name_val).strip().replace("\n", " ") if name_val is not None else ""
            rate = clean_rate(rate_val)
            
            if not code or not name or rate <= 0:
                continue
                
            # Clean up the code to only contain numeric IDs (ignore decimal parts)
            code_clean = code.split('.')[0]
            if not code_clean.isdigit():
                continue
                
            compiled[code_clean] = {
                "id": code_clean,
                "name": name,
                "rate": rate,
                "department": sheetname
            }
            
    print(f"Successfully compiled {len(compiled)} records from {unit_name} template.")
    return compiled

def main():
    folder = r"S:\Sid Work\1. Apollo\@ Apollo Guwahti\Tarriff Working\Tarrif Reporting Format\Tarrif_180826"
    file_intl = os.path.join(folder, "Guwahati assam cash tariff template_International.xlsx")
    file_excl = os.path.join(folder, "Guwahati excel care cash tariff template (1)_Excelcare.xlsx")
    output_path = r"S:\Sid Work\1. Apollo\@ Apollo Guwahti\Tarriff Working\Tarrif Reporting Format\Tarrif Masterss\hdfc_ergo_agreed_2026.json"
    cache_info_path = r"S:\Sid Work\1. Apollo\@ Apollo Guwahti\Tarriff Working\Tarrif Reporting Format\Tarrif Masterss\hdfc_ergo_agreed_cache.json"
    
    intl_hash = get_file_hash(file_intl)
    excl_hash = get_file_hash(file_excl)
    
    # Check cache
    cache_valid = False
    if os.path.exists(cache_info_path) and os.path.exists(output_path):
        try:
            with open(cache_info_path, 'r', encoding='utf-8') as cf:
                cache_data = json.load(cf)
                if cache_data.get("intl_hash") == intl_hash and cache_data.get("excl_hash") == excl_hash:
                    cache_valid = True
        except Exception:
            pass
            
    if cache_valid:
        print("Caching: Centrally agreed HDFC Ergo Excel sheets have not changed. Skipping parsing.")
        return
        
    # Compile both templates
    intl_data = parse_template(file_intl, "International")
    excl_data = parse_template(file_excl, "Excelcare")
    
    # Reconcile side-by-side
    reconciled = []
    all_codes = set(list(intl_data.keys()) + list(excl_data.keys()))
    
    for code in sorted(all_codes):
        intl_item = intl_data.get(code)
        excl_item = excl_data.get(code)
        
        name = intl_item["name"] if intl_item else excl_item["name"]
        dept = intl_item["department"] if intl_item else excl_item["department"]
        
        intl_rate = intl_item["rate"] if intl_item else 0.0
        excl_rate = excl_item["rate"] if excl_item else 0.0
        
        variance = 0.0
        if intl_rate > 0 and excl_rate > 0:
            variance = round(((intl_rate - excl_rate) / excl_rate) * 100, 2)
            
        reconciled.append({
            "id": code,
            "name": name,
            "department": dept,
            "intl_rate": intl_rate,
            "excl_rate": excl_rate,
            "variance": variance
        })
        
    # Write output to json
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(reconciled, f, indent=2, ensure_ascii=False)
    print(f"Saved {len(reconciled)} reconciled records to {output_path}")
    
    # Save cache info
    try:
        with open(cache_info_path, 'w', encoding='utf-8') as cf:
            json.dump({"intl_hash": intl_hash, "excl_hash": excl_hash}, cf, indent=2)
    except Exception as e:
        print(f"Warning: Failed to save cache info: {e}")

if __name__ == "__main__":
    main()
