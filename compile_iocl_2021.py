import openpyxl
import zipfile
import xml.etree.ElementTree as ET
import os
import json
import re
import sys

# Allow overriding paths from command line args
soc_dir = sys.argv[1] if len(sys.argv) > 1 else r"S:\Sid Work\1. Apollo\@ Apollo Guwahti\Tarriff Working\Tarrif Reporting Format\International\Base File\SOC's Use"
output_path = sys.argv[2] if len(sys.argv) > 2 else os.path.join(os.path.dirname(__file__), "compiled_iocl_merged.json")

xlsx_path = os.path.join(soc_dir, "SOC - 2021-22_IOCL.xlsx")
docx_path = os.path.join(soc_dir, "SOC - 2021-22_IOCL.docx")

namespaces = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}

def get_text(tc):
    return " ".join("".join(t.text for t in tc.findall('.//w:t', namespaces) if t.text).strip().split())

def parse_docx_rates(path):
    if not os.path.exists(path):
        print(f"Warning: DOCX file not found at {path}")
        return {}
        
    with zipfile.ZipFile(path) as z:
        doc_xml = z.read('word/document.xml')
        root = ET.fromstring(doc_xml)
        
    tables = root.findall('.//w:tbl', namespaces)
    extracted_data = {} # code -> rates
    
    for t_idx, tbl in enumerate(tables, start=1):
        rows = tbl.findall('.//w:tr', namespaces)
        if not rows:
            continue
        
        header_row_idx = -1
        col_mapping = {}
        
        for r_idx, r in enumerate(rows[:6]):
            cells = [get_text(tc) for tc in r.findall('.//w:tc', namespaces)]
            cells_upper = [c.upper() for c in cells]
            
            has_general = any('GENERAL' == val or 'GENERAL WARD' in val for val in cells_upper)
            has_semi = any('SEMI CABIN' in val or 'SEMI-PRIVATE' in val for val in cells_upper)
            
            if has_general and has_semi:
                header_row_idx = r_idx
                for c_idx, val in enumerate(cells_upper):
                    if any(x in val for x in ['NEW CODE', 'CODE', 'MEDMANTRA CODE', 'HINAI CODE']):
                        col_mapping['code'] = c_idx
                    elif any(x in val for x in ['SERVICE', 'PROCEDURE', 'PARTICULARS', 'NAME OF THE SURGERY']):
                        col_mapping['name'] = c_idx
                    elif 'GENERAL' in val:
                        col_mapping['general'] = c_idx
                    elif 'SEMI CABIN' in val or 'SEMI-PRIVATE' in val:
                        col_mapping['semi'] = c_idx
                    elif any(x in val for x in ['AC CABIN TO SUPER DELUXE', 'AC CABIN', 'CABIN/DELUXE']):
                        col_mapping['ac'] = c_idx
                
                if 'code' not in col_mapping:
                    if len(cells) > 1 and r_idx > 0:
                        prev_cells = [get_text(tc).upper() for tc in rows[r_idx-1].findall('.//w:tc', namespaces)]
                        for c_idx, val in enumerate(prev_cells):
                            if 'NEW CODE' in val or 'CODE' in val:
                                col_mapping['code'] = c_idx
                    else:
                        if len(cells) >= 6:
                            col_mapping['code'] = 1
                            col_mapping['name'] = 2
                break
                
        if header_row_idx == -1 or not ('general' in col_mapping and 'semi' in col_mapping and 'ac' in col_mapping):
            continue
            
        code_col = col_mapping.get('code', 1)
        name_col = col_mapping.get('name', 2)
        gen_col = col_mapping['general']
        semi_col = col_mapping['semi']
        ac_col = col_mapping['ac']
        
        for r_idx in range(header_row_idx + 1, len(rows)):
            cells = [get_text(tc) for tc in rows[r_idx].findall('.//w:tc', namespaces)]
            if len(cells) <= max(code_col, name_col, gen_col, semi_col, ac_col):
                continue
                
            code = cells[code_col].strip()
            name = cells[name_col].strip()
            gen_val = cells[gen_col].strip()
            semi_val = cells[semi_col].strip()
            ac_val = cells[ac_col].strip()
            
            if not code and not name:
                continue
                
            def parse_rate(val):
                if not val or val.lower() in ['no charge', 'free', '-']:
                    return 0.0
                val_clean = re.sub(r'[^\d.]', '', val)
                try:
                    return float(val_clean)
                except ValueError:
                    return None
                    
            r_gen = parse_rate(gen_val)
            r_semi = parse_rate(semi_val)
            r_ac = parse_rate(ac_val)
            
            if r_gen is not None or r_semi is not None or r_ac is not None:
                if code and code.isdigit():
                    extracted_data[code] = {
                        "name": name,
                        "rates": {
                            "GENERAL": r_gen,
                            "SEMI CABIN/ NON AC CABIN": r_semi,
                            "AC CABIN TO SUPER DELUXE AND CRITICAL CARE": r_ac
                        }
                    }
    return extracted_data

SERIAL_NUMBER_SECTIONS = {
    "BED CHARGE",
    "BED CHARGE For Covid - 19",
    "DAY CARE",
    "TRIAGE CHARGE",
    "Indoor consultation fee",
    "TELECONSULTATION CHARGE",
    "INDEX FOR SURGERY",
    "RADIOLOGY INDEX"
}

def compile_iocl_data():
    docx_rates = parse_docx_rates(docx_path)
    
    if not os.path.exists(xlsx_path):
        print(f"Error: XLSX file not found at {xlsx_path}")
        return []
        
    print(f"Loading Excel file from {xlsx_path}...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = wb.active
    
    row_count = sheet.max_row
    print(f"Excel rows: {row_count}")
    
    # 1. Pre-scan parent-row rate patterns
    parent_row_rates = {}
    for r in range(24, row_count + 1):
        c1_val = sheet.cell(r, 1).value
        c1 = str(c1_val).strip() if c1_val is not None else ""
        if c1 and ("Rs." in c1 or "Rs" in c1):
            m = re.search(r'Rs\.?\s*(\d{1,3}(,\d{3})*)', c1)
            if m:
                rate_str = m.group(1).replace(",", "")
                try:
                    rate_num = float(rate_str)
                    name = re.sub(r'Rs\.?.*$', '', c1).strip()
                    
                    for offset in range(1, 5):
                        target_row = r + offset
                        if target_row > row_count:
                            break
                        next_val = sheet.cell(target_row, 1).value
                        if next_val is not None:
                            next_str = str(next_val).strip()
                            if next_str.isdigit() and float(next_str) >= 1:
                                parent_row_rates[next_str] = {"name": name, "rate": rate_num}
                                break
                except Exception:
                    pass

    section = "General"
    list_iocl = []
    
    for r in range(24, row_count + 1):
        c1_val = sheet.cell(r, 1).value
        c4_val = sheet.cell(r, 4).value
        c6_val = sheet.cell(r, 6).value
        c13_val = sheet.cell(r, 13).value
        
        c1 = str(c1_val).strip() if c1_val is not None else ""
        c4 = str(c4_val).strip() if c4_val is not None else ""
        c6 = str(c6_val).strip() if c6_val is not None else ""
        c13 = str(c13_val).strip() if c13_val is not None else ""
        
        # Check section headers
        if c1 and re.match(r'^[A-Za-z /&,-]{5,50}$', c1) and not c6 and not c13:
            section = c1
            continue
        elif c4 and re.match(r'^[A-Za-z /&,-]{5,50}$', c4) and not c1 and not c6 and not c13:
            section = c4
            continue
            
        clean_c1 = ""
        m = re.match(r'^(\d+)', c1)
        if m:
            clean_c1 = m.group(1)
            
        clean_c4 = ""
        m = re.match(r'^(\d+)', c4)
        if m:
            clean_c4 = m.group(1)
            
        code = ""
        code_val = 0
        is_serial_section = (section in SERIAL_NUMBER_SECTIONS)
        
        if clean_c1 and clean_c1.isdigit() and int(clean_c1) >= 1:
            code_val = int(clean_c1)
            if is_serial_section:
                code = f"{section.upper().replace(' ', '_')}_{clean_c1}"
            else:
                code = clean_c1
        elif clean_c4 and clean_c4.isdigit() and int(clean_c4) >= 1:
            code_val = int(clean_c4)
            if is_serial_section:
                code = f"{section.upper().replace(' ', '_')}_{clean_c4}"
            else:
                code = clean_c4
            
        rate = 0.0
        name = ""
        
        if not code:
            rate_col = 0
            for col in range(15, 53):
                v_val = sheet.cell(r, col).value
                v = str(v_val).strip() if v_val is not None else ""
                if v == "No Charge":
                    rate = 0.0
                    rate_col = col
                    break
                elif v:
                    try:
                        rate = float(v)
                        rate_col = col
                        break
                    except ValueError:
                        pass
            if rate_col > 0:
                name_parts = []
                for col in range(1, 15):
                    v_val = sheet.cell(r, col).value
                    v = str(v_val).strip() if v_val is not None else ""
                    if v and len(v) > 1 and not v.isdigit():
                        name_parts.append(v)
                name = " ".join(name_parts)
                if name:
                    code = name
        else:
            if not is_serial_section and str(code_val) in parent_row_rates:
                name = parent_row_rates[str(code_val)]["name"]
                rate = parent_row_rates[str(code_val)]["rate"]
            else:
                rate_col = 0
                for col in range(15, 53):
                    v_val = sheet.cell(r, col).value
                    v = str(v_val).strip() if v_val is not None else ""
                    if v == "No Charge":
                        rate = 0.0
                        rate_col = col
                        break
                    elif v:
                        try:
                            rate = float(v)
                            rate_col = col
                            break
                        except ValueError:
                            pass
                            
                name_parts = []
                start_col = 5 if (clean_c4 and code.endswith(clean_c4)) else 2
                end_col = rate_col - 1 if rate_col > 0 else 18
                if end_col > 18:
                    end_col = 18
                for col in range(start_col, end_col + 1):
                    v_val = sheet.cell(r, col).value
                    v = str(v_val).strip() if v_val is not None else ""
                    if v and len(v) > 1 and not v.isdigit():
                        name_parts.append(v)
                name = " ".join(name_parts)

        if code and name:
            if not is_serial_section and code_val > 0 and code_val < 100 and rate == 0.0 and c1 and sheet.cell(r, 19).value is None:
                continue
                
            record = {
                "id": code,
                "name": name,
                "type": section,
                "dept": section,
                "rate": rate
            }
            
            if code.isdigit() and code in docx_rates:
                record["rates"] = docx_rates[code]["rates"]
                
            list_iocl.append(record)
            
    print(f"Compiled {len(list_iocl)} records.")
    return list_iocl

if __name__ == "__main__":
    records = compile_iocl_data()
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=4)
    print(f"Saved compiled IOCL records to {output_path}")
